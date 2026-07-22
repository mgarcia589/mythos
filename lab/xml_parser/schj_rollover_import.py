"""Schedule J Rollover Analysis + OIT Import Generator.

Compares FY24 ending balances (BalanceBeginningNextYearAmt) against
FY25 beginning balances (BeginningYearBalanceAmt) for all E&P pools.
Generates the OIT Post-86 E&P import file for entities with differences.

Usage:
    python -m lab.xml_parser.schj_rollover_import
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from lab.xml_parser.parser import EFileParser

# --- Config ---
BASE = Path(__file__).resolve().parents[2]
FY24_XML = BASE / "sources" / "xml-parser" / "cb-fy24.xml"
FY25_XML = BASE / "sources" / "xml-parser" / "cb-fy25-v2.xml"
OUTPUT_DIR = BASE / "lab" / "xml_parser" / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# E&P Pools in Schedule J and their descriptions
POOLS = {
    "Post2017EPNotPrevTaxedGrp": "(a) Post-2017 E&P Not Previously Taxed",
    "Section951APTEPGrp": "(c) Section 951A PTEP (GILTI)",
    "Section951a1APTEPGrp": "(e)(i) Section 951(a)(1)(A) PTEP (Sub F)",
    "TotalSection964AEPGrp": "(h) Total Section 964(a) E&P",
    "Post1986UndistributedEarnGrp": "(i) Post-1986 Undistributed Earnings",
    "HoveringDeficitDedSspndTaxGrp": "(j) Hovering Deficit",
}

# OIT Import template columns (65 total)
OIT_COLUMNS = [
    "Entity Number", "Activity Type", "Basket Number", "Basket Name",
    "Tax Year", "Made in Tax Year", "Tax Year Begin", "Tax Year End",
    "Made in Year Begin", "Made in Year End", "Adjustment Made in Current Year",
    "Current Year Adjustment Type", "Current Year Adjustment Type Description",
    "Hovering Deficit Adjustment",
    "959(c)(3) Foreign Source E&P (FC)", "959(c)(3) U.S. Source E&P (FC)",
    "959(c)(3) All Sources Tax (US$)",
    "959(c)(2) OWN Subpart F E&P (FC)", "959(c)(2) OWN 965(a) E&P (FC)",
    "959(c)(2) OWN 965(b) E&P (FC)", "959(c)(2) OWN GILTI E&P (FC)",
    "959(c)(2) OWN 245A(e)(2) E&P (FC)", "959(c)(2) OWN 959(e) E&P (FC)",
    "959(c)(2) OWN 964(e)(4) E&P (FC)",
    "959(c)(2) SUB Subpart F E&P (FC)", "959(c)(2) SUB Subpart F Taxes (US$)",
    "959(c)(2) SUB 965(a) E&P (FC)", "959(c)(2) SUB 965(a) Taxes  (US$)",
    "959(c)(2) SUB 965(b) E&P (FC)", "959(c)(2) SUB 965(b) Taxes  (US$)",
    "959(c)(2) SUB GILTI E&P (FC)", "959(c)(2) SUB GILTI Taxes  (US$)",
    "959(c)(2) SUB 245A(e)(2) E&P (FC)", "959(c)(2) SUB 245A(e)(2) Taxes  (US$)",
    "959(c)(2) SUB 959(e) E&P (FC)", "959(c)(2) SUB 959(e) Taxes  (US$)",
    "959(c)(2) SUB 964(e)(4) E&P (FC)", "959(c)(2) SUB 964(e)(4) Taxes  (US$)",
    "959(c)(1)(B) OWN Excess Passive Assets E&P (FC)",
    "959(c)(1)(B) SUB Excess Passive Assets E&P (FC)",
    "959(c)(1)(B) SUB Excess Passive Assets Taxes  (US$)",
    "959(c)(1)(A) OWN Investment in US Property E&P (FC)",
    "959(c)(1)(A) OWN Subpart F E&P (FC)", "959(c)(1)(A) OWN 965(a) E&P (FC)",
    "959(c)(1)(A) OWN 965(b) E&P (FC)", "959(c)(1)(A) OWN GILTI E&P (FC)",
    "959(c)(1)(A) OWN 245A(e)(2) E&P (FC)", "959(c)(1)(A) OWN 959(e) E&P (FC)",
    "959(c)(1)(A) OWN 964(e)(4) E&P (FC)",
    "959(c)(1)(A) SUB Investment in US Property E&P (FC)",
    "959(c)(1)(A) SUB Investment in US Property Taxes  (US$)",
    "959(c)(1)(A) SUB Subpart F E&P (FC)", "959(c)(1)(A) SUB Subpart F Taxes  (US$)",
    "959(c)(1)(A) SUB 965(a) E&P (FC)", "959(c)(1)(A) SUB 965(a) Taxes  (US$)",
    "959(c)(1)(A) SUB 965(b) E&P (FC)", "959(c)(1)(A) SUB 965(b) Taxes  (US$)",
    "959(c)(1)(A) SUB GILTI E&P (FC)", "959(c)(1)(A) SUB GILTI Taxes  (US$)",
    "959(c)(1)(A) SUB 245A(e)(2) E&P (FC)", "959(c)(1)(A) SUB 245A(e)(2) Taxes  (US$)",
    "959(c)(1)(A) SUB 959(e) E&P (FC)", "959(c)(1)(A) SUB 959(e) Taxes  (US$)",
    "959(c)(1)(A) SUB 964(e)(4) E&P (FC)", "959(c)(1)(A) SUB 964(e)(4) Taxes  (US$)",
]

# Map Sch J pool -> OIT column for the E&P amount
POOL_TO_OIT_COL = {
    "Post2017EPNotPrevTaxedGrp": "959(c)(3) Foreign Source E&P (FC)",
    "Section951APTEPGrp": "959(c)(2) OWN GILTI E&P (FC)",
    "Section951a1APTEPGrp": "959(c)(2) OWN Subpart F E&P (FC)",
    "Post1986UndistributedEarnGrp": None,  # Computed total, not imported directly
    "TotalSection964AEPGrp": None,  # Computed total
    "HoveringDeficitDedSspndTaxGrp": "Hovering Deficit Adjustment",
}


def parse_schedules():
    """Parse Schedule J from both XMLs."""
    fy24 = EFileParser(FY24_XML)
    fy25 = EFileParser(FY25_XML)
    df24 = fy24.extract_form("IRS5471ScheduleJ")
    df25 = fy25.extract_form("IRS5471ScheduleJ")
    return df24, df25


def compare_rollover(df24: pd.DataFrame, df25: pd.DataFrame) -> pd.DataFrame:
    """Compare PY ending vs CY beginning for all pools.

    Handles multi-basket Schedule J data: matches by entity + basket
    (via _basket column if present, or SeparateCategoryCd field).
    """
    results = []

    # Determine basket column
    basket_col_24 = "_basket" if "_basket" in df24.columns else None
    basket_col_25 = "_basket" if "_basket" in df25.columns else None

    def get_basket(row, basket_col):
        if basket_col and row.get(basket_col):
            return str(row[basket_col])
        sep = row.get("IRS5471ScheduleJ_SeparateCategoryCd", "")
        return str(sep) if sep else "UNKNOWN"

    for _, row24 in df24.iterrows():
        ref_id = row24.get("_reference_id", "")
        entity_name = row24.get("_entity_name", "")
        fc = row24.get("_functional_currency", "")
        basket24 = get_basket(row24, basket_col_24)

        if basket24 == "TOTAL":
            continue

        # Match by entity + basket
        match25 = df25[df25["_reference_id"] == ref_id]
        if match25.empty:
            continue

        # Find matching basket in FY25
        row25 = None
        for _, r25 in match25.iterrows():
            if get_basket(r25, basket_col_25) == basket24:
                row25 = r25
                break

        if row25 is None:
            # No matching basket in FY25 — report all pools as missing
            for grp, desc in POOLS.items():
                py_col = f"IRS5471ScheduleJ_{grp}_BalanceBeginningNextYearAmt"
                raw_py = row24.get(py_col, None)
                py_val = pd.to_numeric(raw_py, errors="coerce")
                if pd.isna(py_val):
                    continue
                py_val = float(py_val)
                if abs(py_val) < 1:
                    continue
                results.append({
                    "entity": entity_name,
                    "ref_id": ref_id,
                    "fc": fc,
                    "basket": basket24,
                    "pool": desc,
                    "pool_grp": grp,
                    "py_ending": py_val,
                    "cy_beginning": 0.0,
                    "difference": -py_val,
                    "status": "REVIEW",
                })
            continue

        for grp, desc in POOLS.items():
            py_col = f"IRS5471ScheduleJ_{grp}_BalanceBeginningNextYearAmt"
            cy_col = f"IRS5471ScheduleJ_{grp}_BeginningYearBalanceAmt"

            raw_py = row24.get(py_col, None)
            raw_cy = row25.get(cy_col, None)

            py_val = pd.to_numeric(raw_py, errors="coerce")
            cy_val = pd.to_numeric(raw_cy, errors="coerce")

            if pd.isna(py_val) and pd.isna(cy_val):
                continue

            py_val = 0.0 if pd.isna(py_val) else float(py_val)
            cy_val = 0.0 if pd.isna(cy_val) else float(cy_val)

            diff = cy_val - py_val

            results.append({
                "entity": entity_name,
                "ref_id": ref_id,
                "fc": fc,
                "basket": basket24,
                "pool": desc,
                "pool_grp": grp,
                "py_ending": py_val,
                "cy_beginning": cy_val,
                "difference": diff,
                "status": "OK" if abs(diff) < 10 else "REVIEW",
            })

    return pd.DataFrame(results)


def generate_oit_import(diffs_df: pd.DataFrame) -> pd.DataFrame:
    """Generate OIT import rows for entities with differences.

    Uses the PY ending balance (correct value) as the amount to import,
    since the CY beginning balance is wrong and needs to be overwritten.
    """
    import_rows = []

    entities_with_diffs = diffs_df[diffs_df["status"] == "REVIEW"]

    for ref_id in entities_with_diffs["ref_id"].unique():
        entity_diffs = entities_with_diffs[entities_with_diffs["ref_id"] == ref_id]
        entity_name = entity_diffs.iloc[0]["entity"]

        for _, row in entity_diffs.iterrows():
            pool_grp = row["pool_grp"]
            oit_col = POOL_TO_OIT_COL.get(pool_grp)

            if oit_col is None:
                continue

            import_row = {col: "" for col in OIT_COLUMNS}
            import_row["Entity Number"] = row["ref_id"]
            import_row["Activity Type"] = "EarningsProfitTaxes"
            import_row["Basket Number"] = 21
            import_row["Basket Name"] = "21-General Limitation Income"
            import_row["Tax Year"] = 2024
            import_row["Made in Tax Year"] = 2024
            import_row["Tax Year Begin"] = "01/01/2024"
            import_row["Tax Year End"] = "12/31/2024"
            import_row["Made in Year Begin"] = "01/01/2024"
            import_row["Made in Year End"] = "12/31/2024"
            import_row[oit_col] = row["py_ending"]

            import_rows.append(import_row)

    return pd.DataFrame(import_rows, columns=OIT_COLUMNS)


def write_analysis_report(results_df: pd.DataFrame, output_path: Path):
    """Write the rollover comparison to Excel with formatting."""
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_sum = wb.active
    ws_sum.title = "Summary"

    diffs = results_df[results_df["status"] == "REVIEW"]
    total_checks = len(results_df)
    ok_count = len(results_df[results_df["status"] == "OK"])
    review_count = len(diffs)
    entities_total = results_df["entity"].nunique()
    entities_with_issues = diffs["entity"].nunique()

    header_font = Font(bold=True, size=12)
    ws_sum["A1"] = "SCHEDULE J ROLLOVER ANALYSIS"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A2"] = "Centerbridge Capital Partners III | FY2024 -> FY2025"
    ws_sum["A4"] = "Total Entities Compared:"
    ws_sum["B4"] = entities_total
    ws_sum["A5"] = "Total Pool Checks:"
    ws_sum["B5"] = total_checks
    ws_sum["A6"] = "Clean (OK):"
    ws_sum["B6"] = ok_count
    ws_sum["A7"] = "Differences (REVIEW):"
    ws_sum["B7"] = review_count
    ws_sum["B7"].font = Font(bold=True, color="FF0000") if review_count > 0 else Font()
    ws_sum["A8"] = "Entities with Differences:"
    ws_sum["B8"] = entities_with_issues

    # --- Sheet 2: All Differences ---
    ws_diff = wb.create_sheet("Differences")
    headers = ["Entity", "Ref ID", "FC", "Pool", "PY Ending", "CY Beginning", "Difference", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws_diff.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="D04A02", end_color="D04A02", fill_type="solid")

    for i, (_, row) in enumerate(diffs.iterrows(), 2):
        ws_diff.cell(row=i, column=1, value=row["entity"])
        ws_diff.cell(row=i, column=2, value=row["ref_id"])
        ws_diff.cell(row=i, column=3, value=row["fc"])
        ws_diff.cell(row=i, column=4, value=row["pool"])
        ws_diff.cell(row=i, column=5, value=row["py_ending"])
        ws_diff.cell(row=i, column=6, value=row["cy_beginning"])
        ws_diff.cell(row=i, column=7, value=row["difference"])
        ws_diff.cell(row=i, column=8, value=row["status"])
        ws_diff.cell(row=i, column=5).number_format = "#,##0"
        ws_diff.cell(row=i, column=6).number_format = "#,##0"
        ws_diff.cell(row=i, column=7).number_format = "#,##0"

    # --- Sheet 3: Full Detail (all entities, all pools) ---
    ws_all = wb.create_sheet("Full Detail")
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for i, (_, row) in enumerate(results_df.iterrows(), 2):
        ws_all.cell(row=i, column=1, value=row["entity"])
        ws_all.cell(row=i, column=2, value=row["ref_id"])
        ws_all.cell(row=i, column=3, value=row["fc"])
        ws_all.cell(row=i, column=4, value=row["pool"])
        ws_all.cell(row=i, column=5, value=row["py_ending"])
        ws_all.cell(row=i, column=6, value=row["cy_beginning"])
        ws_all.cell(row=i, column=7, value=row["difference"])
        ws_all.cell(row=i, column=8, value=row["status"])

    wb.save(output_path)
    print(f"Analysis report saved: {output_path}")


def write_oit_import(import_df: pd.DataFrame, output_path: Path):
    """Write OIT-compatible import file."""
    if import_df.empty:
        print("No differences found — no import file needed.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "DE_ImportPost86_EP_Taxes"

    # Write column headers
    for col, h in enumerate(OIT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=9)

    # Write data rows
    for i, (_, row) in enumerate(import_df.iterrows(), 2):
        for col, h in enumerate(OIT_COLUMNS, 1):
            val = row[h]
            if val != "" and val is not None:
                ws.cell(row=i, column=col, value=val)

    wb.save(output_path)
    print(f"OIT Import file saved: {output_path}")


def main():
    print("=" * 70)
    print("SCHEDULE J ROLLOVER: FY2024 ending -> FY2025 beginning")
    print("Centerbridge Capital Partners III")
    print("=" * 70)
    print()

    # Parse
    print("Parsing XMLs...")
    df24, df25 = parse_schedules()
    print(f"  FY24: {len(df24)} entities with Schedule J")
    print(f"  FY25: {len(df25)} entities with Schedule J")
    print()

    # Compare
    print("Comparing rollover balances...")
    results = compare_rollover(df24, df25)

    diffs = results[results["status"] == "REVIEW"]
    ok = results[results["status"] == "OK"]

    print(f"  Total checks: {len(results)}")
    print(f"  OK (match): {len(ok)}")
    print(f"  REVIEW (difference): {len(diffs)}")
    print(f"  Entities with diffs: {diffs['entity'].nunique()}")
    print()

    if not diffs.empty:
        print("MATERIAL DIFFERENCES:")
        print("-" * 70)
        for _, row in diffs.iterrows():
            print(f"  {row['ref_id']:8} | {row['entity'][:30]:30} | {row['pool']}")
            print(f"           PY ending: {row['py_ending']:>15,.0f}  CY beginning: {row['cy_beginning']:>15,.0f}  Diff: {row['difference']:>12,.0f}")
        print()

    # Write analysis report
    report_path = OUTPUT_DIR / "schj-rollover-fy24-fy25.xlsx"
    write_analysis_report(results, report_path)

    # Generate and write OIT import
    print("Generating OIT import file...")
    import_df = generate_oit_import(results)
    if not import_df.empty:
        import_path = OUTPUT_DIR / "oit-import-post86-ep-corrections.xlsx"
        write_oit_import(import_df, import_path)
        print(f"  Import rows generated: {len(import_df)}")
    else:
        print("  No importable differences (totals only or no diffs).")

    print()
    print("Done.")


if __name__ == "__main__":
    main()
