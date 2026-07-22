"""Workbook Reader — Compliance Workbook Ingestion Engine.

Reads PwC 5471/8858 compliance workbooks and extracts structured per-entity
data into DataFrames for reconciliation against IRS e-file XML.

Handles two layout patterns:
  - Layout A (entities-across-columns): Sch H, I-1, I, C, F, E, G
  - Layout B (entities-down-rows): GILTI calc, Subpart F, De Minimis
"""

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.reader import workbook as _wb_module

# Monkey-patch openpyxl to handle invalid print title definitions
_orig_assign = _wb_module.WorkbookParser.assign_names


def _patched_assign(self):
    try:
        _orig_assign(self)
    except (ValueError, Exception):
        pass


_wb_module.WorkbookParser.assign_names = _patched_assign


@dataclass
class EntityInfo:
    entity_code: str
    name: str = ""
    form_type: str = ""  # "5471" or "8858"
    le_id: str = ""
    ein_or_ref: str = ""
    functional_currency: str = ""
    country_code: str = ""
    final_return: bool = False
    dormant: bool = False
    tax_period_begin: str = ""
    tax_period_end: str = ""


@dataclass
class WorkbookData:
    source_path: Path
    client_name: str = ""
    tax_year: str = ""
    entities: dict = field(default_factory=dict)  # entity_code -> EntityInfo
    schedules: dict = field(default_factory=dict)  # schedule_key -> pd.DataFrame

    @property
    def entity_count(self) -> int:
        return len(self.entities)

    @property
    def schedule_keys(self) -> list[str]:
        return sorted(self.schedules.keys())

    def get_schedule(self, key: str) -> pd.DataFrame:
        return self.schedules.get(key, pd.DataFrame())


# Sheet name patterns for schedule detection
SCHEDULE_SHEET_MAP = {
    "sch_h_fc": ["5471 Sch H - FC", "5471 Sch H"],
    "sch_i1_fc": ["5471 Sch I-1 - FC", "5471 Sch I-1"],
    "sch_i1_usd": ["5471 Sch I-1 - USD"],
    "sch_i": ["5471 Sch I"],
    "sch_c_fc": ["5471 Sch C - FC", "C - FC"],
    "sch_c_usd": ["5471 Sch C - USD", "C - USD"],
    "sch_f_usd": ["5471 Sch F - USD", "5471 Sch F"],
    "sch_e": ["5471 Sch E"],
    "sch_g": ["5471 Sch G"],
    "gilti_calc": ["2025 GILTI", "2024 GILTI", "GILTI"],
    "subf_summary": ["Subpart F - Summary"],
    "de_minimis": ["De Minimis - Analysis", "De Minimis"],
    "entity_listing": ["Entity Listing"],
}

# Layout A: row label structure for each schedule
# Maps line_number -> (row_offset_from_data_start, xml_field_name)
LAYOUT_A_SCHEDULES = {
    "sch_h_fc": {
        "header_row": 4,  # Entity Code row
        "data_start_row": 11,  # First data row (line 1)
        "data_start_col": 5,  # Column E (1-indexed)
        "lines": [
            (0, "1", "ForeignCYNetIncomePerBooksAmt", "CY Net Income per Books"),
            (2, "2a", "CapitalGainsOrLossesAmt", "Capital gains or losses"),
            (3, "2b", "DepreciationAndAmortizationAmt", "Depreciation and amortization"),
            (4, "2c", "DepletionAmt", "Depletion"),
            (5, "2d", "InvestmentOrIncentiveAllwncAmt", "Investment or incentive allowance"),
            (6, "2e", "ChargesToStatutoryReservesAmt", "Charges to statutory reserves"),
            (7, "2f", "InventoryAdjustmentsAmt", "Inventory adjustments"),
            (8, "2g", "TaxesNetAddnAmt", "Income taxes"),
            (9, "2h", "FrgnCurrencyGainLossAddnAmt", "Foreign currency gains/losses"),
            (10, "2i_sub", "OtherAdjustmentsNetSbtrctnAmt", "Other net subtractions"),
            (11, "2i_add", "OtherAdjustmentsNetAddnAmt", "Other net additions"),
            (13, "3", "TotalNetAdditionsAmt", "Total net additions"),
            (14, "4", "TotalNetSubtractionsAmt", "Total net subtractions"),
            (15, "5a", "CurrentEarningsAndProfitsAmt", "Current E&P"),
            (16, "5b", "DASTMGainOrLossAmt", "DASTM gain or loss"),
            (17, "5c", "EarningAndPrftPlusDASTMGainAmt", "Current E&P after DASTM"),
            (18, "5d", "CurrEarnAndPrftInUSDollarsAmt", "Current E&P in USD"),
        ],
    },
    "sch_i1_fc": {
        "header_row": 4,
        "data_start_row": 9,
        "data_start_col": 3,  # Column C
        "lines": [
            (0, "1", "GrossIncomeAmt", "Gross income"),
            (1, "2a", "ExclGrossIncmEffCntdFCCorpAmt", "ECI exclusion"),
            (2, "2b", "ExclGrossIncmSubpartFIncmAmt", "Subpart F exclusion"),
            (3, "2c", "ExclGrossIncmHghTxdIncmAmt", "High-tax exclusion"),
            (4, "2d", "ExclGrossIncmDvdRcvdAmt", "Dividends received exclusion"),
            (5, "2e", "ExclGrossIncmFrgnOilGasAmt", "Foreign oil/gas exclusion"),
            (6, "3", "TotalExclusionsAmt", "Total exclusions"),
            (7, "4", "GrossIncmLessExclusionsAmt", "Gross income less exclusions"),
            (8, "5", "AllocableDedExpnssAmt", "Allocable deductions"),
            (9, "6", "TestedIncomeLossAmt", "Tested income / (loss)"),
            (10, "7", "TestedForeignIncomeTaxesAmt", "Tested foreign income taxes"),
            (11, "8", "QBAIAmt", "QBAI"),
        ],
    },
    "sch_i": {
        "header_row": 4,
        "data_start_row": 9,
        "data_start_col": 3,
        "lines": [
            (0, "1a", "Sect964e4SubpartFDvdAmt", "Section 964(e)(4) dividends"),
            (1, "1b", "Sect245Ae2SubpartFIncmAmt", "Section 245A(e)(2) income"),
            (2, "1c", "TieredExtraordinaryDisposAmt", "Tiered extraordinary dispositions"),
            (3, "1d", "TieredExtraordinaryReductAmt", "Tiered extraordinary reductions"),
            (4, "1e", "SubpartFPHCIncomeAmt", "FPHCI"),
            (5, "1f", "FBCSalesIncomeAmt", "FBC Sales income"),
            (6, "1g", "FBCServicesIncomeAmt", "FBC Services income"),
            (7, "1h", "OtherSubpartFIncomeAmt", "Other subpart F income"),
            (8, "2", "EarningsInvestedUSPropAmt", "Earnings invested in US property"),
            (10, "4", "FactoringIncomeAmt", "Factoring income"),
            (11, "5a", "Sect245AEligibleDividendsAmt", "Section 245A eligible dividends"),
            (12, "5b", "ExtraordinaryDispositionAmt", "Extraordinary disposition amounts"),
        ],
    },
    "sch_c_fc": {
        "header_row": 5,
        "data_start_row": 11,
        "data_start_col": 3,
        "lines": [
            (0, "1a", "ForeignGrossReceiptsOrSalesAmt", "Gross receipts or sales"),
            (1, "1b", "ForeignReturnsAndAllowancesAmt", "Returns and allowances"),
            (2, "1c", "ForeignNetReceiptsAmt", "Net receipts"),
            (3, "2", "ForeignCostOfGoodsSoldAmt", "Cost of goods sold"),
            (4, "3", "ForeignGrossProfitAmt", "Gross profit"),
            (5, "4", "ForeignDividendIncomeAmt", "Dividends"),
            (6, "5", "ForeignInterestIncomeAmt", "Interest"),
            (7, "6a", "ForeignGrossRentsAmt", "Gross rents"),
            (8, "6b", "ForeignGrossRoyaltiesAmt", "Gross royalties and license fees"),
            (9, "7", "ForeignNetGainLossAmt", "Net gain/loss on sale of capital assets"),
            (10, "8a", "ForeignCurrencyUnrealizedAmt", "FX gain/loss — unrealized"),
            (11, "8b", "ForeignCurrencyRealizedAmt", "FX gain/loss — realized"),
            (12, "9", "ForeignOtherIncomeAmt", "Other income"),
            (13, "10", "ForeignTotalIncomeAmt", "Total income"),
            (14, "11", "ForeignCompensationNotDeductedAmt", "Compensation not deducted"),
            (15, "12a", "ForeignRentsDeductionAmt", "Rents"),
            (16, "12b", "ForeignRoyaltiesDeductionAmt", "Royalties and license fees"),
            (17, "13", "ForeignInterestExpenseAmt", "Interest expense"),
            (18, "14", "ForeignDepreciationNotDeductedAmt", "Depreciation not deducted"),
            (19, "15", "ForeignDepletionAmt", "Depletion"),
            (20, "16", "ForeignTaxesAmt", "Taxes (excl income tax)"),
            (21, "17", "ForeignOtherDeductionsAmt", "Other deductions"),
            (22, "18", "ForeignTotalDeductionsAmt", "Total deductions"),
            (23, "19", "ForeignNetIncomeBeforeTaxAmt", "Net income before tax"),
            (24, "20", "ForeignUnusualItemsAmt", "Unusual or infrequently occurring items"),
            (25, "21a", "ForeignIncomeTaxCurrentAmt", "Income tax — current"),
            (26, "21b", "ForeignIncomeTaxDeferredAmt", "Income tax — deferred"),
            (27, "22", "ForeignCYNetIncomePerBookAmt", "Current year net income per books"),
            (28, "23a", "ForeignCurrencyTranslationAmt", "FX translation adjustments"),
            (29, "23b", "ForeignOtherComprehensiveAmt", "Other comprehensive income"),
            (30, "23c", "ForeignIncomeTaxOCIAmt", "Income tax expense on OCI"),
            (31, "24", "ForeignTotalComprehensiveAmt", "Other comprehensive income net of tax"),
        ],
    },
    "sch_c_usd": {
        "header_row": 5,
        "data_start_row": 11,
        "data_start_col": 3,
        "lines": [
            (0, "1a", "ForeignGrossReceiptsOrSalesAmt", "Gross receipts or sales"),
            (1, "1b", "ForeignReturnsAndAllowancesAmt", "Returns and allowances"),
            (2, "1c", "ForeignNetReceiptsAmt", "Net receipts"),
            (3, "2", "ForeignCostOfGoodsSoldAmt", "Cost of goods sold"),
            (4, "3", "ForeignGrossProfitAmt", "Gross profit"),
            (5, "4", "ForeignDividendIncomeAmt", "Dividends"),
            (6, "5", "ForeignInterestIncomeAmt", "Interest"),
            (7, "6a", "ForeignGrossRentsAmt", "Gross rents"),
            (8, "6b", "ForeignGrossRoyaltiesAmt", "Gross royalties and license fees"),
            (9, "7", "ForeignNetGainLossAmt", "Net gain/loss on sale of capital assets"),
            (10, "8a", "ForeignCurrencyUnrealizedAmt", "FX gain/loss — unrealized"),
            (11, "8b", "ForeignCurrencyRealizedAmt", "FX gain/loss — realized"),
            (12, "9", "ForeignOtherIncomeAmt", "Other income"),
            (13, "10", "ForeignTotalIncomeAmt", "Total income"),
            (14, "11", "ForeignCompensationNotDeductedAmt", "Compensation not deducted"),
            (15, "12a", "ForeignRentsDeductionAmt", "Rents"),
            (16, "12b", "ForeignRoyaltiesDeductionAmt", "Royalties and license fees"),
            (17, "13", "ForeignInterestExpenseAmt", "Interest expense"),
            (18, "14", "ForeignDepreciationNotDeductedAmt", "Depreciation not deducted"),
            (19, "15", "ForeignDepletionAmt", "Depletion"),
            (20, "16", "ForeignTaxesAmt", "Taxes (excl income tax)"),
            (21, "17", "ForeignOtherDeductionsAmt", "Other deductions"),
            (22, "18", "ForeignTotalDeductionsAmt", "Total deductions"),
            (23, "19", "ForeignNetIncomeBeforeTaxAmt", "Net income before tax"),
            (24, "20", "ForeignUnusualItemsAmt", "Unusual or infrequently occurring items"),
            (25, "21a", "ForeignIncomeTaxCurrentAmt", "Income tax — current"),
            (26, "21b", "ForeignIncomeTaxDeferredAmt", "Income tax — deferred"),
            (27, "22", "ForeignCYNetIncomePerBookAmt", "Current year net income per books"),
            (28, "23a", "ForeignCurrencyTranslationAmt", "FX translation adjustments"),
            (29, "23b", "ForeignOtherComprehensiveAmt", "Other comprehensive income"),
            (30, "23c", "ForeignIncomeTaxOCIAmt", "Income tax expense on OCI"),
            (31, "24", "ForeignTotalComprehensiveAmt", "Other comprehensive income net of tax"),
        ],
    },
    "sch_f_usd": {
        "header_row": 5,
        "data_start_row": 11,
        "data_start_col": 3,
        "lines": [
            # Assets
            (0, "1", "EndAcctPrdCashAmt", "Cash"),
            (1, "2a", "EndAcctPrdTradeNotesAmt", "Trade notes & accounts receivable"),
            (2, "2b", "EndAcctPrdBadDebtAllowanceAmt", "Less allowance for bad debts"),
            (3, "3", "EndAcctPrdDerivativesAmt", "Derivatives"),
            (4, "4", "EndAcctPrdInventoriesAmt", "Inventories"),
            (5, "5", "EndAcctPrdOtherCurrentAmt", "Other current assets"),
            (6, "6", "EndAcctPrdLoansToShareholdersAmt", "Loans to shareholders"),
            (7, "7", "EndAcctPrdInvstSubsidiaryAmt", "Investment in subsidiaries"),
            (8, "8", "EndAcctPrdOtherInvestmentsAmt", "Other investments"),
            (9, "9a", "EndAcctPrdBldgAndOtherAstAmt", "Buildings & depreciable assets"),
            (10, "9b", "EndAcctPrdAccumDepreciationAmt", "Less accumulated depreciation"),
            (11, "10a", "EndAcctPrdDepletableAssetsAmt", "Depletable assets"),
            (12, "10b", "EndAcctPrdAccumDepletionAmt", "Less accumulated depletion"),
            (13, "11", "EndAcctPrdLandAmt", "Land"),
            (14, "12", "EndAcctPrdIntangibleAssetsAmt", "Intangible assets"),
            (15, "12a", "EndAcctPrdGoodwillAmt", "Goodwill"),
            (16, "12b", "EndAcctPrdOrgCostsAmt", "Organization costs"),
            (17, "12c", "EndAcctPrdPatentsOthAstAmt", "Patents, trademarks, other intangibles"),
            (18, "12d", "EndAcctPrdAccumAmortizationAmt", "Less accumulated amortization"),
            (19, "13", "EndAcctPrdOtherAssetsAmt", "Other assets"),
            (20, "14", "EndAcctPrdTotalAssetsAmt", "Total assets"),
            # Liabilities (offset from row 11: rows 40-51 = offsets 29-40)
            (29, "15", "EndAcctPrdAccountsPayableAmt", "Accounts payable"),
            (30, "16", "EndAcctPrdOtherCurrLiabAmt", "Other current liabilities"),
            (31, "17", "EndAcctPrdLiabDerivativesAmt", "Derivatives (liability)"),
            (32, "18", "EndAcctPrdLoansFromShareholdersAmt", "Loans from shareholders"),
            (33, "19", "EndAcctPrdOthLiabilitiesAmt", "Other liabilities"),
            (34, "20", "EndAcctPrdCapitalStockAmt", "Capital stock"),
            (35, "20a", "EndAcctPrdPreferredStockAmt", "Preferred stock"),
            (36, "20b", "EndAcctPrdCommonStockAmt", "Common stock"),
            (37, "21", "EndAcctPrdPaidInOrSurplusAmt", "Paid-in capital or surplus"),
            (38, "22", "EndAcctPrdRtnEarningsAmt", "Retained earnings"),
            (39, "23", "EndAcctPrdTreasuryStockAmt", "Less cost of treasury stock"),
            (40, "24", "EndAcctPrdTotLiabShrEqtyAmt", "Total liabilities & shareholders equity"),
        ],
    },
    "sch_e": {
        "header_row": 5,
        "data_start_row": 9,
        "data_start_col": 3,
        "lines": [
            (0, "d", "SchECountryCodeTxt", "Country code"),
            (1, "e", "SchEForeignTaxYearTxt", "Foreign tax year"),
            (2, "f", "SchEUSTaxYearTxt", "US tax year"),
            (3, "g", "SchEIncomeSubjectToTaxAmt", "Income subject to tax"),
            (4, "h", "SchETaxesPaidUSSourceInd", "Taxes paid on US source income?"),
            (5, "i", "SchELocalCurrencyTxt", "Local currency"),
            (6, "j", "TotalTaxInFunctionalCurAmt", "Tax paid/accrued in local currency"),
            (7, "k", "SchETranslationRateRt", "Translation rate to USD"),
            (8, "l", "TotalTaxInUSDollarsAmt", "Amount in USD"),
            (9, "m", "SchETaxInFunctionalCurAmt", "Amount in functional currency"),
        ],
    },
    "sch_g": {
        "header_row": 6,
        "data_start_row": 10,
        "data_start_col": 7,
        "lines": [
            (0, "1", "SchGLine1Ind", "10% interest in foreign partnership?"),
            (1, "2", "SchGLine2Ind", "Interest in any trust?"),
            (2, "3", "SchGLine3Ind", "Disregarded entities or branches?"),
            (3, "4A", "BaseErosionPaymentBenefitInd", "Base erosion payments (BEAT)?"),
            (4, "4B", "SchGLine4BAmt", "Total base erosion payment amount"),
            (5, "4C", "SchGLine4CAmt", "Total base erosion tax benefit amount"),
            (6, "5A", "SchGLine5AInd", "163(j) limitation election?"),
            (7, "5B", "SchGLine5BAmt", "163(j) limited amount"),
            (8, "6A", "SchGLine6AInd", "BEAT applies?"),
            (9, "6B", "SchGLine6BAmt", "BEAT amount"),
            (10, "6C", "SchGLine6CAmt", "BEAT adjusted gross income"),
            (11, "6D", "SchGLine6DAmt", "BEAT tax amount"),
            (12, "7", "SchGLine7Ind", "Dual consolidated loss?"),
            (13, "8", "SchGLine8Ind", "Extraordinary reduction?"),
            (14, "9A", "SchGLine9AInd", "Extraordinary disposition?"),
            (15, "9B", "SchGLine9BAmt", "Extraordinary disposition amount"),
            (16, "10", "SchGLine10Ind", "Section 965 specified foreign corp?"),
            (17, "11", "SchGLine11Ind", "Section 245A shareholder?"),
            (18, "12", "SchGLine12Ind", "GILTI high-tax exclusion election?"),
            (19, "13", "SchGLine13Ind", "Subpart F high-tax exception election?"),
        ],
    },
}


class WorkbookReader:
    """Read PwC compliance workbooks and extract structured data."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")
        self._wb = openpyxl.load_workbook(
            str(self.path), read_only=True, data_only=True
        )
        self._sheet_names = self._wb.sheetnames

    def close(self):
        self._wb.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    @property
    def sheet_names(self) -> list[str]:
        return self._sheet_names

    def _find_sheet(self, schedule_key: str) -> Optional[str]:
        """Find the actual sheet name for a given schedule key."""
        candidates = SCHEDULE_SHEET_MAP.get(schedule_key, [])
        for candidate in candidates:
            if candidate in self._sheet_names:
                return candidate
        return None

    def parse(self) -> WorkbookData:
        """Full extraction — parse all available schedules."""
        data = WorkbookData(source_path=self.path)

        # Extract metadata from first available schedule sheet
        self._extract_metadata(data)

        # Entity listing
        data.entities = self._parse_entity_listing()

        # Layout A schedules (entities across columns)
        for key in ["sch_h_fc", "sch_i1_fc", "sch_i", "sch_c_fc", "sch_c_usd",
                    "sch_f_usd", "sch_e", "sch_g"]:
            df = self._parse_layout_a(key)
            if not df.empty:
                data.schedules[key] = df

        # Layout B schedules (entities down rows)
        for key in ["gilti_calc", "subf_summary", "de_minimis"]:
            df = self._parse_layout_b(key)
            if not df.empty:
                data.schedules[key] = df

        return data

    def _extract_metadata(self, data: WorkbookData):
        """Extract client name and tax year from workbook header rows."""
        for key in ["sch_h_fc", "sch_i1_fc", "entity_listing"]:
            sheet_name = self._find_sheet(key)
            if not sheet_name:
                continue
            ws = self._wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, max_row=3, max_col=2, values_only=True))
            if rows:
                data.client_name = str(rows[0][0] or "")
                for row in rows[1:]:
                    val = str(row[0] or "")
                    if "FY" in val.upper() or "20" in val:
                        # Extract year from strings like "FY2025 Compliance"
                        for word in val.split():
                            if word.startswith("FY") and len(word) >= 4:
                                data.tax_year = word[2:]
                                break
                            elif word.startswith("20") and len(word) == 4:
                                data.tax_year = word
                                break
            break

    def _parse_entity_listing(self) -> dict[str, EntityInfo]:
        """Parse Entity Listing sheet into entity registry."""
        sheet_name = self._find_sheet("entity_listing")
        if not sheet_name:
            return {}

        ws = self._wb[sheet_name]
        entities = {}

        for row in ws.iter_rows(min_row=3, max_col=12, values_only=True):
            entity_id = str(row[2] or "").strip() if row[2] else ""
            entity_name = str(row[3] or "").strip() if row[3] else ""

            if not entity_id or not entity_name:
                continue

            entities[entity_id] = EntityInfo(
                entity_code=entity_id,
                name=entity_name,
                form_type=str(row[4] or ""),
                le_id=str(row[1] or ""),
                ein_or_ref=str(row[7] or ""),
                functional_currency=str(row[8] or ""),
                country_code=str(row[9] or ""),
                final_return=str(row[5] or "").lower() == "yes",
                dormant=str(row[6] or "").lower() == "yes",
                tax_period_begin=str(row[10] or "")[:10],
                tax_period_end=str(row[11] or "")[:10],
            )

        return entities

    def _parse_layout_a(self, schedule_key: str) -> pd.DataFrame:
        """Parse a Layout A sheet (entities across columns)."""
        sheet_name = self._find_sheet(schedule_key)
        if not sheet_name:
            return pd.DataFrame()

        config = LAYOUT_A_SCHEDULES.get(schedule_key)
        if not config:
            return pd.DataFrame()

        ws = self._wb[sheet_name]

        # Read entity codes from header row
        header_row = list(ws.iter_rows(
            min_row=config["header_row"],
            max_row=config["header_row"],
            values_only=True
        ))[0]

        data_start_col = config["data_start_col"]
        entity_codes = []
        for i in range(data_start_col - 1, len(header_row)):
            code = str(header_row[i] or "").strip()
            if code and code.startswith("C"):
                entity_codes.append((i, code))
            elif code and not code.startswith("C"):
                # Sometimes header has entity name instead of code — stop
                break

        if not entity_codes:
            return pd.DataFrame()

        # Read all data rows at once
        max_line_offset = max(line[0] for line in config["lines"])
        data_rows = list(ws.iter_rows(
            min_row=config["data_start_row"],
            max_row=config["data_start_row"] + max_line_offset,
            values_only=True
        ))

        # Build records
        records = {}
        for col_idx, entity_code in entity_codes:
            record = {"_entity_code": entity_code}

            for offset, line_num, field_name, description in config["lines"]:
                if offset < len(data_rows):
                    row_data = data_rows[offset]
                    if col_idx < len(row_data):
                        value = row_data[col_idx]
                        record[field_name] = value if value is not None else 0
                    else:
                        record[field_name] = 0
                else:
                    record[field_name] = 0

            records[entity_code] = record

        df = pd.DataFrame.from_dict(records, orient="index")
        df.index.name = "_entity_code"
        return df

    def _parse_layout_b(self, schedule_key: str) -> pd.DataFrame:
        """Parse a Layout B sheet (entities down rows)."""
        sheet_name = self._find_sheet(schedule_key)
        if not sheet_name:
            return pd.DataFrame()

        ws = self._wb[sheet_name]

        if schedule_key == "gilti_calc":
            return self._parse_gilti_calc(ws)
        elif schedule_key == "subf_summary":
            return self._parse_subf_summary(ws)
        elif schedule_key == "de_minimis":
            return self._parse_de_minimis(ws)

        return pd.DataFrame()

    def _parse_gilti_calc(self, ws) -> pd.DataFrame:
        """Parse 2025 GILTI sheet."""
        # Header is row 7, data starts row 8+
        header_row = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]

        # Build column map from header
        col_names = []
        for v in header_row:
            col_names.append(str(v).strip() if v else "")

        # Find entity number column (typically col 3)
        entity_col_idx = None
        for i, name in enumerate(col_names):
            if "entity number" in name.lower() or "entity code" in name.lower():
                entity_col_idx = i
                break

        if entity_col_idx is None:
            entity_col_idx = 2  # default: column C

        records = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            entity_code = str(row[entity_col_idx] or "").strip() if row[entity_col_idx] else ""
            if not entity_code or not entity_code.startswith("C"):
                continue

            record = {"_entity_code": entity_code}

            # Map known columns by header name
            col_map = {
                "FC": "functional_currency",
                "Entity Name": "entity_name",
                "Type of Entity": "entity_type",
                "Book Income": "book_income",
                "E&P Adjustments": "ep_adjustments",
                "163(j) Disallowance / (CF)": "s163j_disallowance",
                "Net Subpart F Pre FI": "net_subf_pre_fi",
                "Full Inclusion - Other Sub F": "full_inclusion_other_subf",
                "Tested Income": "tested_income_fc",
                "FX": "fx_rate",
                "Tested Loss (USD)": "tested_loss_usd",
                "QBAI": "qbai",
                "Interest Expense": "interest_expense",
                "Interest Income": "interest_income",
                "Specified Interest Expense": "specified_interest_expense",
                "Foreign Taxes": "foreign_taxes",
            }

            for i, header in enumerate(col_names):
                mapped = col_map.get(header)
                if mapped and i < len(row):
                    record[mapped] = row[i] if row[i] is not None else 0

            # Tested Income (USD) — there are two columns, take the first non-empty
            for i, header in enumerate(col_names):
                if "tested income" in header.lower() and "usd" in header.lower():
                    if i < len(row) and row[i]:
                        record["tested_income_usd"] = row[i]
                        break

            records.append(record)

        if not records:
            return pd.DataFrame()

        df = pd.DataFrame(records)
        df = df.set_index("_entity_code")
        return df

    def _parse_subf_summary(self, ws) -> pd.DataFrame:
        """Parse Subpart F - Summary sheet."""
        # Header is row 5, data starts row 7
        header_row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]

        col_names = [str(v).strip() if v else "" for v in header_row]

        records = []
        for row in ws.iter_rows(min_row=7, values_only=True):
            # Entity Code in col B (index 1)
            entity_code = str(row[1] or "").strip() if len(row) > 1 and row[1] else ""
            if not entity_code or not entity_code.startswith("C"):
                continue

            record = {
                "_entity_code": entity_code,
                "entity_name": str(row[2] or "") if len(row) > 2 else "",
                "pre_tax_subpart_f": row[3] if len(row) > 3 else 0,
                "tax_on_subf": row[4] if len(row) > 4 else 0,
                "etr_on_subf": row[5] if len(row) > 5 else 0,
                "eligible_high_tax": str(row[6] or "") if len(row) > 6 else "",
                "net_subpart_f": row[7] if len(row) > 7 else 0,
            }
            records.append(record)

        if not records:
            return pd.DataFrame()

        df = pd.DataFrame(records)
        df = df.set_index("_entity_code")
        return df

    def _parse_de_minimis(self, ws) -> pd.DataFrame:
        """Parse De Minimis - Analysis sheet."""
        # Header is row 6, data starts row 8
        header_row = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
        col_names = [str(v).strip() if v else "" for v in header_row]

        records = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            entity_code = str(row[0] or "").strip() if row[0] else ""
            if not entity_code or not entity_code.startswith("C"):
                continue

            record = {
                "_entity_code": entity_code,
                "entity_name": str(row[1] or "") if len(row) > 1 else "",
                "interest_income": row[2] if len(row) > 2 else 0,
                "dividend_income": row[3] if len(row) > 3 else 0,
                "fx_gain_loss": row[4] if len(row) > 4 else 0,
                "passive_fphci": row[5] if len(row) > 5 else 0,
                "fphci_ic_interest": row[6] if len(row) > 6 else 0,
                "us_source_fphci": row[7] if len(row) > 7 else 0,
            }
            records.append(record)

        if not records:
            return pd.DataFrame()

        df = pd.DataFrame(records)
        df = df.set_index("_entity_code")
        return df

    def extract_schedule(self, schedule_key: str) -> pd.DataFrame:
        """Extract a single schedule by key."""
        if schedule_key in LAYOUT_A_SCHEDULES:
            return self._parse_layout_a(schedule_key)
        return self._parse_layout_b(schedule_key)

    def detect_profile(self) -> str:
        """Auto-detect workbook profile based on sheet names."""
        names = set(self._sheet_names)
        if "Entity Listing" in names and any("5471 Sch" in s for s in names):
            return "centerbridge"
        if "Foreign Summary" in names and "951A Summary" in names:
            return "cng_ep_gilti"
        if "5471 Sch M" in names:
            return "cng_sch_m"
        if "Distributions" in names and "Sec. 986(c)" in " ".join(names):
            return "cng_dividends"
        return "unknown"

    def parse_auto(self) -> WorkbookData:
        """Auto-detect profile and parse accordingly."""
        profile = self.detect_profile()
        if profile == "centerbridge":
            return self.parse()
        elif profile == "cng_ep_gilti":
            return self._parse_cng_ep_gilti()
        return WorkbookData(source_path=self.path)

    def _parse_cng_ep_gilti(self) -> WorkbookData:
        """Parse CNG-style E&P/GILTI/SubF workbook (Foreign Summary sheet)."""
        data = WorkbookData(source_path=self.path)

        ws = self._wb["Foreign Summary"]

        # Metadata from rows 1-4
        row1 = list(ws.iter_rows(min_row=1, max_row=1, max_col=2, values_only=True))[0]
        data.client_name = str(row1[0] or "")
        row2 = list(ws.iter_rows(min_row=2, max_row=2, max_col=2, values_only=True))[0]
        label = str(row2[0] or "")
        for word in label.split():
            if word.startswith("20") and len(word) == 4:
                data.tax_year = word
                break

        # Entity codes from row 5
        row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
        entity_codes = []
        for i, v in enumerate(row5):
            code = str(v or "").strip()
            if code and "FC" in code.upper():
                entity_codes.append((i, code))

        if not entity_codes:
            return data

        # Build entity registry from codes + row 6 (names) + row 8 (FC)
        row6 = list(ws.iter_rows(min_row=6, max_row=6, values_only=True))[0]
        row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
        for col_idx, code in entity_codes:
            name = str(row6[col_idx]) if col_idx < len(row6) and row6[col_idx] else ""
            fc = str(row8[col_idx]) if col_idx < len(row8) and row8[col_idx] else ""
            data.entities[code] = EntityInfo(
                entity_code=code, name=name, form_type="5471",
                functional_currency=fc,
            )

        # Extract key row data into a schedule DataFrame
        # These rows map to reconcilable XML fields
        ROW_MAP = {
            13: ("pbt_usd", "Pre-tax book income (USD)"),
            15: ("net_income_usd", "Net income (USD)"),
            73: ("total_ep", "Total E&P"),
            76: ("pretax_ep", "Pre-Tax E&P"),
            80: ("subpart_f", "Subpart F income"),
            82: ("dividends", "Dividends"),
            87: ("tested_income_loss", "Tested Income/(Loss) @100%"),
            88: ("tested_foreign_taxes", "Tested Foreign Income Taxes"),
            89: ("pretax_tested_income", "Pre-tax Tentative Tested Income"),
            100: ("tested_income_after_hte", "Tested Income After HTE"),
            103: ("qbai", "QBAI"),
            107: ("interest_income", "Interest Income"),
            112: ("interest_expense", "Interest Expense"),
            118: ("income_taxes", "Income Taxes"),
        }

        records = {}
        for col_idx, code in entity_codes:
            record = {"_entity_code": code}
            for row_num, (field_name, _) in ROW_MAP.items():
                val = ws.cell(row_num, col_idx + 1).value  # openpyxl is 1-indexed
                record[field_name] = val if val is not None else 0
            records[code] = record

        df = pd.DataFrame.from_dict(records, orient="index")
        df.index.name = "_entity_code"
        data.schedules["foreign_summary"] = df

        # Sch H reconcilable fields
        # WB row 73 "Total E&P" is USD-translated → maps to XML CurrEarnAndPrftInUSDollarsAmt
        # WB row 15 "Net Income" is a WB-specific calc (no direct XML match at Sch H)
        ep_records = {}
        for col_idx, code in entity_codes:
            ep_records[code] = {
                "_entity_code": code,
                "CurrEarnAndPrftInUSDollarsAmt": ws.cell(73, col_idx + 1).value or 0,
            }

        df_h = pd.DataFrame.from_dict(ep_records, orient="index")
        df_h.index.name = "_entity_code"
        data.schedules["sch_h_fc"] = df_h

        # GILTI data — Sch I-1 equivalents
        # WB row 100 "Tested Income After HTE" = XML TestedIncomeLossGrp_USDollarAmt
        #   HTE entities: WB=0, XML=NaN (DORMANT_OK pattern)
        #   Non-HTE entities: WB amount ≈ XML amount (within rounding)
        # WB row 87 "Tested @100% pre-HTE" — no clean XML equivalent (gross income is different concept)
        gilti_records = {}
        for col_idx, code in entity_codes:
            gilti_records[code] = {
                "_entity_code": code,
                "TestedIncomeLossGrp_USDollarAmt": ws.cell(100, col_idx + 1).value or 0,
            }

        df_i1 = pd.DataFrame.from_dict(gilti_records, orient="index")
        df_i1.index.name = "_entity_code"
        data.schedules["sch_i1_fc"] = df_i1

        return data

    def summary(self) -> str:
        """Print a quick summary of what's available."""
        data = self.parse_auto()
        lines = [
            f"Workbook: {self.path.name}",
            f"Profile: {self.detect_profile()}",
            f"Client: {data.client_name}",
            f"Tax Year: {data.tax_year}",
            f"Entities: {data.entity_count}",
            f"Schedules parsed: {len(data.schedules)}",
        ]
        for key, df in data.schedules.items():
            lines.append(f"  {key}: {len(df)} entities x {len(df.columns)} fields")
        return "\n".join(lines)
