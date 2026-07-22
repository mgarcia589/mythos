"""Schedule C & F: XML vs Workbook Comparison Tool.

Compares IRS 5471 Schedule C (Income Statement) and Schedule F (Balance Sheet)
from OIT XML against the engagement workbook for Centerbridge FY2025.

Usage:
    python -m lab.xml_parser.schcf_compare <xml_path> <workbook_path>

Or programmatically:
    from lab.xml_parser.schcf_compare import compare_schc, compare_schf
"""

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from lab.xml_parser.parser import EFileParser

# --- Config ---
BASE = Path(__file__).resolve().parents[2]
OUTPUT_DIR = BASE / "lab" / "xml_parser" / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Schedule C line mapping: XML field -> (Line #, Description)
SCHC_LINES_FC = {
    "ForeignGrossReceiptsOrSalesAmt": ("1a", "Gross receipts or sales"),
    "ForeignReturnsAndAllowancesAmt": ("1b", "Returns and allowances"),
    "ForeignNetGrossReceiptsAmt": ("1c", "Net gross receipts (1a - 1b)"),
    "ForeignCostOfGoodsSoldAmt": ("2", "Cost of goods sold"),
    "ForeignGrossProfitAmt": ("3", "Gross profit (1c - 2)"),
    "ForeignInterestIncomeAmt": ("4", "Interest income"),
    "ForeignGrossRentsAmt": ("5", "Gross rents"),
    "ForeignRentsAmt": ("6", "Royalties/rents"),
    "FrgnCurTransGainLossrlzdAmt": ("7", "FX gain/loss realized"),
    "ForeignOtherIncomeAmt": ("8", "Other income"),
    "ForeignTotalIncomeAmt": ("9", "Total income (3 through 8)"),
    "ForeignCompensationNotDedAmt": ("10", "Compensation not deducted"),
    "ForeignInterestDeductionAmt": ("11", "Interest deductions"),
    "ForeignDepreciationNotDedAmt": ("12", "Depreciation not deducted"),
    "ForeignTaxesAmt": ("13", "Taxes"),
    "ForeignOtherDeductionsAmt": ("14", "Other deductions"),
    "ForeignTotalDeductionsAmt": ("15", "Total deductions (10 through 14)"),
    "FrgnTotalIncomeMinusTotDedAmt": ("16", "Net income (9 - 15)"),
    "FrgnCurrentIncomeTaxExpenseAmt": ("17a", "Current income tax expense"),
    "FrgnDefrdIncomeTaxExpenseAmt": ("17b", "Deferred income tax expense"),
    "ForeignCYNetIncomePerBooksAmt": ("18", "CY net income per books"),
}

SCHC_LINES_USD = {
    "USGrossReceiptsOrSalesAmt": ("1a", "Gross receipts or sales"),
    "USReturnsAndAllowancesAmt": ("1b", "Returns and allowances"),
    "USNetGrossReceiptsAmt": ("1c", "Net gross receipts (1a - 1b)"),
    "USCostOfGoodsSoldAmt": ("2", "Cost of goods sold"),
    "USGrossProfitAmt": ("3", "Gross profit (1c - 2)"),
    "USInterestIncomeAmt": ("4", "Interest income"),
    "USGrossRentsAmt": ("5", "Gross rents"),
    "USRentsAmt": ("6", "Royalties/rents"),
    "USCurTransGainLossRlzdAmt": ("7", "FX gain/loss realized"),
    "USOtherIncomeAmt": ("8", "Other income"),
    "USTotalIncomeAmt": ("9", "Total income (3 through 8)"),
    "USCompensationNotDedAmt": ("10", "Compensation not deducted"),
    "USInterestDeductionAmt": ("11", "Interest deductions"),
    "USDepreciationNotDedAmt": ("12", "Depreciation not deducted"),
    "USTaxesAmt": ("13", "Taxes"),
    "USOtherDeductionsAmt": ("14", "Other deductions"),
    "USTotalDeductionsAmt": ("15", "Total deductions (10 through 14)"),
    "USTotalIncomeMinusTotDedAmt": ("16", "Net income (9 - 15)"),
    "USCurrentIncomeTaxExpenseAmt": ("17a", "Current income tax expense"),
    "USDefrdIncomeTaxExpenseAmt": ("17b", "Deferred income tax expense"),
    "USCYNetIncomePerBooksAmt": ("18", "CY net income per books"),
}

# Schedule F line mapping: XML field -> (Line #, Description)
SCHF_LINES_BEGIN = {
    "BegngAcctPrdCashAmt": ("1", "Cash"),
    "BegngAcctPrdTradeNotesAmt": ("2", "Trade notes & accounts receivable"),
    "BegngAcctPrdBadDebtAllwncAmt": ("2a", "Less: bad debt allowance"),
    "BegngAcctPrdInventoriesAmt": ("3", "Inventories"),
    "BegngAcctPrdOtherCurrAssetsAmt": ("4", "Other current assets"),
    "BegngAcctPrdLoansToShrAmt": ("5", "Loans to shareholders"),
    "BegngAcctPrdInvstSubsidiaryAmt": ("6", "Investment in subsidiaries"),
    "BegngAcctPrdOthInvestmentsAmt": ("7", "Other investments"),
    "BegngAcctPrdBldgAndOtherAstAmt": ("8a", "Buildings & other depreciable assets"),
    "BegngAcctPrdNetAccumDeprecAmt": ("8b", "Less: accumulated depreciation"),
    "BegngAcctPrdLandAmt": ("9a", "Land"),
    "BegngAcctPrdNetAccumAmortzAmt": ("9b", "Less: accumulated amortization"),
    "BegngAcctPrdPatentsOthAstAmt": ("10a", "Patents/intangibles/other assets"),
    "BegngAcctPrdGoodwillAmt": ("10b", "Goodwill"),
    "BegngAcctPrdOtherAssetsAmt": ("11", "Other assets"),
    "BegngAcctPrdTotalAssetsAmt": ("12", "TOTAL ASSETS"),
    "BegngAcctPrdAccountsPayableAmt": ("13", "Accounts payable"),
    "BegngAcctPrdOtherCurrLiabAmt": ("14", "Other current liabilities"),
    "BegngAcctPrdLoansFromShrAmt": ("15", "Loans from shareholders"),
    "BegngAcctPrdOthLiabilitiesAmt": ("16", "Other liabilities"),
    "BegngAcctPrdCommonStockAmt": ("17", "Capital stock"),
    "BegngAcctPrdPaidInOrSurplusAmt": ("18", "Paid-in or capital surplus"),
    "BegngAcctPrdRtnEarningsAmt": ("19", "Retained earnings"),
    "BegngAcctPrdTotLiabShrEqtyAmt": ("20", "TOTAL LIABILITIES & EQUITY"),
}

SCHF_LINES_END = {
    "EndAcctPrdCashAmt": ("1", "Cash"),
    "EndAcctPrdTradeNotesAmt": ("2", "Trade notes & accounts receivable"),
    "EndAcctPrdBadDebtAllwncAmt": ("2a", "Less: bad debt allowance"),
    "EndAcctPrdInventoriesAmt": ("3", "Inventories"),
    "EndAcctPrdOtherCurrAssetsAmt": ("4", "Other current assets"),
    "EndAcctPrdLoansToShrAmt": ("5", "Loans to shareholders"),
    "EndAcctPrdInvstSubsidiaryAmt": ("6", "Investment in subsidiaries"),
    "EndAcctPrdOthInvestmentsAmt": ("7", "Other investments"),
    "EndAcctPrdBldgAndOtherAstAmt": ("8a", "Buildings & other depreciable assets"),
    "EndAcctPrdNetAccumDeprecAmt": ("8b", "Less: accumulated depreciation"),
    "EndAcctPrdLandAmt": ("9a", "Land"),
    "EndAcctPrdNetAccumAmortzAmt": ("9b", "Less: accumulated amortization"),
    "EndAcctPrdPatentsOthAstAmt": ("10a", "Patents/intangibles/other assets"),
    "EndAcctPrdGoodwillAmt": ("10b", "Goodwill"),
    "EndAcctPrdOtherAssetsAmt": ("11", "Other assets"),
    "EndAcctPrdTotalAssetsAmt": ("12", "TOTAL ASSETS"),
    "EndAcctPrdAccountsPayableAmt": ("13", "Accounts payable"),
    "EndAcctPrdOtherCurrLiabAmt": ("14", "Other current liabilities"),
    "EndAcctPrdLoansFromShrAmt": ("15", "Loans from shareholders"),
    "EndAcctPrdOthLiabilitiesAmt": ("16", "Other liabilities"),
    "EndAcctPrdCommonStockAmt": ("17", "Capital stock"),
    "EndAcctPrdPaidInOrSurplusAmt": ("18", "Paid-in or capital surplus"),
    "EndAcctPrdRtnEarningsAmt": ("19", "Retained earnings"),
    "EndAcctPrdTotLiabShrEqtyAmt": ("20", "TOTAL LIABILITIES & EQUITY"),
}


def extract_xml_schc(xml_path: str | Path) -> pd.DataFrame:
    """Extract Schedule C data from XML, one row per entity."""
    parser = EFileParser(xml_path)
    df = parser.extract_form("IRS5471")

    records = []
    for _, row in df.iterrows():
        entity = row.get("_entity_name", "")
        ref_id = row.get("_reference_id", "")
        fc = row.get("_functional_currency", "")

        record = {
            "entity_name": entity,
            "ref_id": ref_id,
            "functional_currency": fc,
        }

        # FC columns
        for field, (line, desc) in SCHC_LINES_FC.items():
            col = "IRS5471_IRS5471ScheduleC_" + field
            val = pd.to_numeric(row.get(col, None), errors="coerce")
            record["FC_L" + line] = val if not pd.isna(val) else None

        # USD columns
        for field, (line, desc) in SCHC_LINES_USD.items():
            col = "IRS5471_IRS5471ScheduleC_" + field
            val = pd.to_numeric(row.get(col, None), errors="coerce")
            record["USD_L" + line] = val if not pd.isna(val) else None

        records.append(record)

    return pd.DataFrame(records)


def extract_xml_schf(xml_path: str | Path) -> pd.DataFrame:
    """Extract Schedule F data from XML, one row per entity."""
    parser = EFileParser(xml_path)
    df = parser.extract_form("IRS5471")

    records = []
    for _, row in df.iterrows():
        entity = row.get("_entity_name", "")
        ref_id = row.get("_reference_id", "")
        fc = row.get("_functional_currency", "")

        record = {
            "entity_name": entity,
            "ref_id": ref_id,
            "functional_currency": fc,
        }

        # Beginning of year
        for field, (line, desc) in SCHF_LINES_BEGIN.items():
            col = "IRS5471_IRS5471ScheduleF_" + field
            val = pd.to_numeric(row.get(col, None), errors="coerce")
            record["BOY_L" + line] = val if not pd.isna(val) else None

        # End of year
        for field, (line, desc) in SCHF_LINES_END.items():
            col = "IRS5471_IRS5471ScheduleF_" + field
            val = pd.to_numeric(row.get(col, None), errors="coerce")
            record["EOY_L" + line] = val if not pd.isna(val) else None

        records.append(record)

    return pd.DataFrame(records)


def compare_schc(xml_df: pd.DataFrame, wb_df: pd.DataFrame, threshold: float = 1.0) -> pd.DataFrame:
    """Compare Schedule C XML vs workbook.

    Args:
        xml_df: Output of extract_xml_schc()
        wb_df: Workbook data with same structure (entity_name/ref_id + FC_L*/USD_L* columns)
        threshold: Minimum absolute difference to flag

    Returns:
        DataFrame with differences
    """
    results = []
    compare_cols = [c for c in xml_df.columns if c.startswith(("FC_L", "USD_L"))]

    for _, xml_row in xml_df.iterrows():
        ref_id = xml_row["ref_id"]
        entity = xml_row["entity_name"]

        # Match by ref_id
        wb_match = wb_df[wb_df["ref_id"] == ref_id]
        if wb_match.empty:
            results.append({
                "entity": entity,
                "ref_id": ref_id,
                "line": "ALL",
                "description": "Entity in XML but NOT in workbook",
                "xml_value": None,
                "wb_value": None,
                "difference": None,
                "status": "MISSING_WB",
            })
            continue

        wb_row = wb_match.iloc[0]
        for col in compare_cols:
            xml_val = xml_row.get(col)
            wb_val = wb_row.get(col)

            if pd.isna(xml_val) and pd.isna(wb_val):
                continue
            xml_val = 0.0 if pd.isna(xml_val) else float(xml_val)
            wb_val = 0.0 if pd.isna(wb_val) else float(wb_val)

            diff = xml_val - wb_val
            if abs(diff) >= threshold:
                currency = "FC" if col.startswith("FC") else "USD"
                line = col.split("_L")[1]
                line_info = SCHC_LINES_FC if currency == "FC" else SCHC_LINES_USD
                desc = ""
                for field, (ln, d) in line_info.items():
                    if ln == line:
                        desc = d
                        break

                results.append({
                    "entity": entity,
                    "ref_id": ref_id,
                    "line": line,
                    "currency": currency,
                    "description": desc,
                    "xml_value": xml_val,
                    "wb_value": wb_val,
                    "difference": diff,
                    "status": "DIFF",
                })

    return pd.DataFrame(results)


def compare_schf(xml_df: pd.DataFrame, wb_df: pd.DataFrame, threshold: float = 1.0) -> pd.DataFrame:
    """Compare Schedule F XML vs workbook.

    Args:
        xml_df: Output of extract_xml_schf()
        wb_df: Workbook data with same structure (entity_name/ref_id + BOY_L*/EOY_L* columns)
        threshold: Minimum absolute difference to flag

    Returns:
        DataFrame with differences
    """
    results = []
    compare_cols = [c for c in xml_df.columns if c.startswith(("BOY_L", "EOY_L"))]

    for _, xml_row in xml_df.iterrows():
        ref_id = xml_row["ref_id"]
        entity = xml_row["entity_name"]

        wb_match = wb_df[wb_df["ref_id"] == ref_id]
        if wb_match.empty:
            results.append({
                "entity": entity,
                "ref_id": ref_id,
                "line": "ALL",
                "description": "Entity in XML but NOT in workbook",
                "xml_value": None,
                "wb_value": None,
                "difference": None,
                "status": "MISSING_WB",
            })
            continue

        wb_row = wb_match.iloc[0]
        for col in compare_cols:
            xml_val = xml_row.get(col)
            wb_val = wb_row.get(col)

            if pd.isna(xml_val) and pd.isna(wb_val):
                continue
            xml_val = 0.0 if pd.isna(xml_val) else float(xml_val)
            wb_val = 0.0 if pd.isna(wb_val) else float(wb_val)

            diff = xml_val - wb_val
            if abs(diff) >= threshold:
                period = "BOY" if col.startswith("BOY") else "EOY"
                line = col.split("_L")[1]
                line_info = SCHF_LINES_BEGIN if period == "BOY" else SCHF_LINES_END
                desc = ""
                for field, (ln, d) in line_info.items():
                    if ln == line:
                        desc = d
                        break

                results.append({
                    "entity": entity,
                    "ref_id": ref_id,
                    "line": line,
                    "period": period,
                    "description": desc,
                    "xml_value": xml_val,
                    "wb_value": wb_val,
                    "difference": diff,
                    "status": "DIFF",
                })

    return pd.DataFrame(results)


def write_comparison_report(schc_diffs: pd.DataFrame, schf_diffs: pd.DataFrame, output_path: Path):
    """Write comparison results to formatted Excel."""
    wb = Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "SCHEDULE C & F: XML vs WORKBOOK COMPARISON"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Centerbridge Capital Partners III | FY2025"
    ws["A4"] = "Schedule C Differences:"
    ws["B4"] = len(schc_diffs[schc_diffs["status"] == "DIFF"]) if not schc_diffs.empty else 0
    ws["A5"] = "Schedule C Missing from WB:"
    ws["B5"] = len(schc_diffs[schc_diffs["status"] == "MISSING_WB"]) if not schc_diffs.empty else 0
    ws["A6"] = "Schedule F Differences:"
    ws["B6"] = len(schf_diffs[schf_diffs["status"] == "DIFF"]) if not schf_diffs.empty else 0
    ws["A7"] = "Schedule F Missing from WB:"
    ws["B7"] = len(schf_diffs[schf_diffs["status"] == "MISSING_WB"]) if not schf_diffs.empty else 0

    header_fill = PatternFill(start_color="D04A02", end_color="D04A02", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Schedule C sheet
    if not schc_diffs.empty:
        ws_c = wb.create_sheet("Sch C Diffs")
        headers = list(schc_diffs.columns)
        for col, h in enumerate(headers, 1):
            cell = ws_c.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, (_, row) in enumerate(schc_diffs.iterrows(), 2):
            for col, h in enumerate(headers, 1):
                ws_c.cell(row=i, column=col, value=row[h])

    # Schedule F sheet
    if not schf_diffs.empty:
        ws_f = wb.create_sheet("Sch F Diffs")
        headers = list(schf_diffs.columns)
        for col, h in enumerate(headers, 1):
            cell = ws_f.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, (_, row) in enumerate(schf_diffs.iterrows(), 2):
            for col, h in enumerate(headers, 1):
                ws_f.cell(row=i, column=col, value=row[h])

    wb.save(output_path)
    print("Report saved: %s" % output_path)


if __name__ == "__main__":
    print("=" * 70)
    print("SCHEDULE C & F: XML vs WORKBOOK COMPARISON")
    print("Centerbridge Capital Partners III | FY2025")
    print("=" * 70)
    print()
    print("XML extraction ready. Waiting for workbook path to run comparison.")
    print()
    print("Usage:")
    print("  1. Pass workbook path as argument, or")
    print("  2. Call extract_xml_schc() / extract_xml_schf() programmatically")
    print()

    # Quick preview of XML data
    xml_path = BASE / "sources" / "xml-parser" / "cb-fy25-v2.xml"
    print("Extracting Schedule C from XML...")
    schc = extract_xml_schc(xml_path)
    print("  Entities with Sch C data: %d" % len(schc[schc.filter(like="FC_L").notna().any(axis=1)]))
    print()

    print("Extracting Schedule F from XML...")
    schf = extract_xml_schf(xml_path)
    print("  Entities with Sch F data: %d" % len(schf[schf.filter(like="BOY_L").notna().any(axis=1)]))
    print()
    print("Ready for workbook comparison. Provide WB data to compare_schc() / compare_schf().")
